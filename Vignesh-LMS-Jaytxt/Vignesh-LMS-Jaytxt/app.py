import os
import sys
import calendar
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from dotenv import load_dotenv
from sqlalchemy.orm import joinedload
from functools import wraps
from reportlab.lib.units import inch
# Load environment variables
load_dotenv()

app = Flask(__name__)

# Get the directory where this script is located
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# SIMPLE CONFIGURATION - Using current directory
app.config['SECRET_KEY'] = 'dev-secret-key-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(BASE_DIR, "textile_lms.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(BASE_DIR, 'static', 'uploads')
app.config['ANNUAL_LEAVE_DAYS'] = 12
app.config['SICK_LEAVE_DAYS'] = 10
app.config['CASUAL_LEAVE_DAYS'] = 7
app.config['EMERGENCY_LEAVE_DAYS'] = 5

# Print debug info
print(f"🔧 Base directory: {BASE_DIR}")
print(f"🔧 Database path: {app.config['SQLALCHEMY_DATABASE_URI']}")

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'


# Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(20), unique=True, nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone = db.Column(db.String(20))
    department = db.Column(db.String(50))
    designation = db.Column(db.String(50))
    shift = db.Column(db.String(20))
    date_of_joining = db.Column(db.Date, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    # Relationships
    leaves = db.relationship('Leave', foreign_keys='Leave.user_id', backref='applicant', lazy=True)
    attendances = db.relationship('Attendance', foreign_keys='Attendance.user_id', backref='employee', lazy=True)
    approved_leaves = db.relationship('Leave', foreign_keys='Leave.approved_by', backref='approver', lazy=True)
    recorded_attendances = db.relationship('Attendance', foreign_keys='Attendance.recorded_by', backref='recorder', lazy=True)

    @property
    def password(self):
        raise AttributeError('password is not a readable attribute')

    @password.setter
    def password(self, password):
        self.password_hash = generate_password_hash(password)

    def verify_password(self, password):
        return check_password_hash(self.password_hash, password)

    def get_full_name(self):
        return f"{self.first_name} {self.last_name}"


class Leave(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    leave_type = db.Column(db.String(50), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=False)
    total_days = db.Column(db.Integer, nullable=False)
    reason = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='Pending')
    admin_comment = db.Column(db.Text)
    applied_date = db.Column(db.DateTime, default=datetime.utcnow)
    approved_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    approved_date = db.Column(db.DateTime, nullable=True)


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    check_in = db.Column(db.Time, nullable=True)
    check_out = db.Column(db.Time, nullable=True)
    status = db.Column(db.String(20), default='Absent')
    overtime_hours = db.Column(db.Float, default=0)
    remarks = db.Column(db.Text, nullable=True)
    recorded_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)


# Decorator for admin-only routes
def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('user_dashboard'))
        return f(*args, **kwargs)

    return decorated_function


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# Routes
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        employee_id = request.form.get('employee_id')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        phone = request.form.get('phone')
        department = request.form.get('department')
        designation = request.form.get('designation')
        date_of_joining = request.form.get('date_of_joining')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            flash('Passwords do not match!', 'danger')
            return redirect(url_for('register'))

        if User.query.filter_by(email=email).first():
            flash('Email already registered!', 'danger')
            return redirect(url_for('register'))

        if User.query.filter_by(employee_id=employee_id).first():
            flash('Employee ID already exists!', 'danger')
            return redirect(url_for('register'))

        try:
            user = User(
                employee_id=employee_id,
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                department=department,
                designation=designation,
                date_of_joining=datetime.strptime(date_of_joining, '%Y-%m-%d').date(),
                is_admin=False,
                is_active=True
            )
            user.password = password
            db.session.add(user)
            db.session.commit()

            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Registration failed: {str(e)}', 'danger')
            return redirect(url_for('register'))

    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = True if request.form.get('remember') else False

        user = User.query.filter_by(email=email).first()

        if user and user.verify_password(password):
            if not user.is_active:
                flash('Account is deactivated. Please contact HR.', 'danger')
                return redirect(url_for('login'))

            login_user(user, remember=remember)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password!', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))


@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_admin:
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('user_dashboard'))


@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    # Dashboard statistics
    total_employees = User.query.filter_by(is_admin=False, is_active=True).count()
    pending_leaves = Leave.query.filter_by(status='Pending').count()
    today_attendance = Attendance.query.filter_by(date=date.today(), status='Present').count()

    # Simple query without joins - we'll handle names separately
    recent_leaves = Leave.query.order_by(Leave.applied_date.desc()).limit(10).all()

    # Get user details for each leave
    leaves_with_users = []
    for leave in recent_leaves:
        user = User.query.get(leave.user_id)
        leaves_with_users.append({
            'leave': leave,
            'user': user
        })

    # Attendance overview for the week
    week_attendance = []
    for i in range(7):
        day = date.today() - timedelta(days=i)
        present_count = Attendance.query.filter_by(date=day, status='Present').count()
        week_attendance.append({
            'date': day.strftime('%Y-%m-%d'),
            'present': present_count
        })

    return render_template('admin/dashboard.html',
                           total_employees=total_employees,
                           pending_leaves=pending_leaves,
                           today_attendance=today_attendance,
                           leaves_with_users=leaves_with_users,
                           week_attendance=week_attendance)

@app.route('/user/dashboard')
@login_required
def user_dashboard():
    # User statistics
    total_leaves = Leave.query.filter_by(user_id=current_user.id).count()
    approved_leaves = Leave.query.filter_by(user_id=current_user.id, status='Approved').count()
    pending_leaves = Leave.query.filter_by(user_id=current_user.id, status='Pending').count()

    # Today's attendance
    today_attendance = Attendance.query.filter_by(
        user_id=current_user.id,
        date=date.today()
    ).first()

    # Recent leaves
    recent_leaves = Leave.query.filter_by(user_id=current_user.id) \
        .order_by(Leave.applied_date.desc()).limit(5).all()

    # Leave balances
    annual_taken = Leave.query.filter_by(
        user_id=current_user.id,
        leave_type='Annual',
        status='Approved'
    ).count()
    sick_taken = Leave.query.filter_by(
        user_id=current_user.id,
        leave_type='Sick',
        status='Approved'
    ).count()

    annual_balance = app.config['ANNUAL_LEAVE_DAYS'] - annual_taken
    sick_balance = app.config['SICK_LEAVE_DAYS'] - sick_taken

    return render_template('user/dashboard.html',
                           total_leaves=total_leaves,
                           approved_leaves=approved_leaves,
                           pending_leaves=pending_leaves,
                           today_attendance=today_attendance,
                           recent_leaves=recent_leaves,
                           annual_balance=annual_balance,
                           sick_balance=sick_balance)


@app.route('/apply_leave', methods=['GET', 'POST'])
@login_required
def apply_leave():
    if request.method == 'POST':
        leave_type = request.form.get('leave_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        reason = request.form.get('reason')
        coworker_id = request.form.get('coworker_id')
        is_new_worker = request.form.get('is_new_worker') == '1'

        # Determine who the leave is for
        if is_new_worker:
            # Apply leave for a new worker (create user first)
            new_first_name = request.form.get('new_first_name')
            new_last_name = request.form.get('new_last_name')
            new_employee_id = request.form.get('new_employee_id')
            new_department = request.form.get('new_department')
            new_designation = request.form.get('new_designation')
            new_phone = request.form.get('new_phone')
            new_email = request.form.get('new_email')
            new_ticket_number = request.form.get('new_worker_ticket_number')

            # Check if employee ID already exists
            existing_user = User.query.filter_by(employee_id=new_employee_id).first()
            if existing_user:
                flash(f'Employee ID {new_employee_id} already exists!', 'danger')
                return redirect(url_for('apply_leave'))

            # Create a temporary email if not provided
            if not new_email or new_email.strip() == '':
                new_email = f'{new_employee_id.lower()}@temp.textile.com'

            # Create temporary password
            temp_password = 'TempPass123'

            # Create new user
            new_user = User(
                employee_id=new_employee_id,
                first_name=new_first_name,
                last_name=new_last_name,
                email=new_email,
                phone=new_phone,
                department=new_department,
                designation=new_designation,
                date_of_joining=date.today(),
                is_admin=False,
                is_active=True  # Or set to False if you want HR to activate manually
            )
            new_user.password = temp_password

            try:
                db.session.add(new_user)
                db.session.flush()  # Get the ID without committing

                user_id = new_user.id

                # Store ticket number in a new field or use a separate table
                # You might want to add a ticket_number field to the Leave model
                ticket_number = new_ticket_number

                flash(f'New worker {new_first_name} {new_last_name} created successfully!', 'success')

            except Exception as e:
                db.session.rollback()
                flash(f'Error creating new worker: {str(e)}', 'danger')
                return redirect(url_for('apply_leave'))

        elif coworker_id and coworker_id != '':
            # Apply leave for existing coworker
            user_id = int(coworker_id)
            is_coworker_application = True
            ticket_number = request.form.get('ticket_number')
        else:
            # Apply leave for self
            user_id = current_user.id
            is_coworker_application = False
            ticket_number = None

        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        # Calculate total days excluding weekends
        delta = (end - start).days + 1
        total_days = 0
        for i in range(delta):
            day = start + timedelta(days=i)
            if day.weekday() < 5:  # Monday to Friday
                total_days += 1

        # Check leave balance for self applications only
        if user_id == current_user.id:
            if leave_type == 'Annual':
                annual_taken = Leave.query.filter_by(
                    user_id=current_user.id,
                    leave_type='Annual',
                    status='Approved'
                ).count()
                annual_balance = app.config['ANNUAL_LEAVE_DAYS'] - annual_taken
                if total_days > annual_balance:
                    flash(f'Insufficient Annual leave balance! You have {annual_balance} days left.', 'danger')
                    return redirect(url_for('apply_leave'))

            elif leave_type == 'Sick':
                sick_taken = Leave.query.filter_by(
                    user_id=current_user.id,
                    leave_type='Sick',
                    status='Approved'
                ).count()
                sick_balance = app.config['SICK_LEAVE_DAYS'] - sick_taken
                if total_days > sick_balance:
                    flash(f'Insufficient Sick leave balance! You have {sick_balance} days left.', 'danger')
                    return redirect(url_for('apply_leave'))

        # Create leave application
        leave = Leave(
            user_id=user_id,
            leave_type=leave_type,
            start_date=start,
            end_date=end,
            total_days=total_days,
            reason=reason
        )

        # If you added applied_by field to Leave model, set it for coworker applications
        # if user_id != current_user.id:
        #     leave.applied_by = current_user.id

        # If you added ticket_number field to Leave model
        # if ticket_number:
        #     leave.ticket_number = ticket_number

        db.session.add(leave)
        db.session.commit()

        # Show appropriate success message
        if is_new_worker:
            flash(f'Leave application submitted successfully for new worker {new_first_name} {new_last_name}!',
                  'success')
        elif user_id != current_user.id:
            coworker = User.query.get(user_id)
            flash(f'Leave application submitted successfully for {coworker.get_full_name()}!', 'success')
        else:
            flash('Leave application submitted successfully!', 'success')

        return redirect(url_for('leave_status'))

    # For GET request, get coworkers list for the dropdown
    coworkers = User.query.filter(
        User.is_admin == False,
        User.is_active == True,
        User.id != current_user.id
    ).order_by(User.first_name, User.last_name).all()

    # Calculate leave balances for self
    annual_taken = Leave.query.filter_by(
        user_id=current_user.id,
        leave_type='Annual',
        status='Approved'
    ).count()
    sick_taken = Leave.query.filter_by(
        user_id=current_user.id,
        leave_type='Sick',
        status='Approved'
    ).count()

    casual_taken = Leave.query.filter_by(
        user_id=current_user.id,
        leave_type='Casual',
        status='Approved'
    ).count()
    emergency_taken = Leave.query.filter_by(
        user_id=current_user.id,
        leave_type='Emergency',
        status='Approved'
    ).count()

    annual_balance = app.config['ANNUAL_LEAVE_DAYS'] - annual_taken
    sick_balance = app.config['SICK_LEAVE_DAYS'] - sick_taken
    casual_balance = app.config['CASUAL_LEAVE_DAYS'] - casual_taken
    emergency_balance = app.config['EMERGENCY_LEAVE_DAYS'] - emergency_taken

    # Get recent leaves for the current user
    recent_leaves = Leave.query.filter_by(user_id=current_user.id) \
        .order_by(Leave.applied_date.desc()).limit(5).all()

    return render_template('user/apply_leave.html',
                           coworkers=coworkers,
                           annual_balance=annual_balance,
                           sick_balance=sick_balance,
                           casual_balance=casual_balance,
                           emergency_balance=emergency_balance,
                           recent_leaves=recent_leaves)


@app.route('/user/leave_status')
@login_required
def leave_status():
    # Get all leaves for current user
    leaves = Leave.query.filter_by(user_id=current_user.id).order_by(Leave.start_date.desc()).all()

    # Calculate statistics
    total_leaves = len(leaves)
    approved_leaves = len([l for l in leaves if l.status == 'Approved'])
    pending_leaves = len([l for l in leaves if l.status == 'Pending'])
    rejected_leaves = len([l for l in leaves if l.status == 'Rejected'])

    return render_template('user/leave_status.html',
                           leaves=leaves,
                           total_leaves=total_leaves,
                           approved_leaves=approved_leaves,
                           pending_leaves=pending_leaves,
                           rejected_leaves=rejected_leaves)


from flask import make_response, request, send_file
import csv
import io
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO


@app.route('/admin/leaves/export', methods=['POST'])
@login_required
@admin_required
def export_leaves():
    format_type = request.form.get('format', 'csv')
    status_filter = request.form.get('status', 'all')
    leave_type_filter = request.form.get('leave_type')
    department_filter = request.form.get('department')
    date_from_filter = request.form.get('date_from')
    search_filter = request.form.get('search')

    # Build query with filters
    query = Leave.query.join(User, Leave.user_id == User.id)  # Fixed: user_id instead of applicant_id

    if status_filter != 'all':
        query = query.filter(Leave.status == status_filter)

    if leave_type_filter:
        query = query.filter(Leave.leave_type == leave_type_filter)

    if department_filter:
        query = query.filter(User.department == department_filter)

    if date_from_filter:
        query = query.filter(Leave.start_date >= datetime.strptime(date_from_filter, '%Y-%m-%d').date())

    if search_filter:
        query = query.filter(
            db.or_(
                User.first_name.ilike(f'%{search_filter}%'),
                User.last_name.ilike(f'%{search_filter}%'),
                User.employee_id.ilike(f'%{search_filter}%')
            )
        )

    leaves = query.order_by(Leave.applied_date.desc()).all()

    if format_type == 'csv':
        return export_leaves_csv(leaves)
    elif format_type == 'pdf':
        return export_leaves_pdf(leaves)
    else:
        return "Invalid format", 400

def export_leaves_csv(leaves):
    """Export leaves as CSV"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['Employee ID', 'Employee Name', 'Department', 'Leave Type',
                     'Start Date', 'End Date', 'Total Days', 'Reason',
                     'Applied Date', 'Status', 'Approved/Rejected By', 'Admin Comment'])

    # Write data
    for leave in leaves:
        writer.writerow([
            leave.applicant.employee_id,
            leave.applicant.get_full_name(),
            leave.applicant.department,
            leave.leave_type,
            leave.start_date.strftime('%Y-%m-%d'),
            leave.end_date.strftime('%Y-%m-%d'),
            leave.total_days,
            leave.reason,
            leave.applied_date.strftime('%Y-%m-%d'),
            leave.status,
            leave.approver.get_full_name() if leave.approver else '',
            leave.admin_comment or ''
        ])

    # Create response
    response = make_response(output.getvalue())
    response.headers[
        'Content-Disposition'] = f'attachment; filename=leaves_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response.headers['Content-type'] = 'text/csv'
    return response


def export_leaves_pdf(leaves):
    """Export leaves as PDF"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors

        # Create PDF in memory
        pdf_buffer = BytesIO()

        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(f"<b>Leave Requests Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Create table data
        table_data = [['Employee', 'Department', 'Leave Type', 'From', 'To', 'Days', 'Status']]

        for leave in leaves:
            table_data.append([
                leave.applicant.get_full_name(),
                leave.applicant.department or '',
                leave.leave_type,
                leave.start_date.strftime('%Y-%m-%d'),
                leave.end_date.strftime('%Y-%m-%d'),
                str(leave.total_days),
                leave.status
            ])

        # Create table
        table = Table(table_data, colWidths=[2 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1 * inch, 0.5 * inch, 1 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        pdf_buffer.seek(0)

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'leaves_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )

    except ImportError:
        # If reportlab is not installed, create a simple text file
        text_content = "Leave Requests Report\n"
        text_content += f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"

        for leave in leaves:
            text_content += f"{leave.applicant.get_full_name()} | {leave.applicant.department} | {leave.leave_type} | "
            text_content += f"{leave.start_date.strftime('%Y-%m-%d')} to {leave.end_date.strftime('%Y-%m-%d')} | "
            text_content += f"{leave.total_days} days | {leave.status}\n"

        response = make_response(text_content)
        response.headers[
            'Content-Disposition'] = f'attachment; filename=leaves_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        response.headers['Content-type'] = 'text/plain'
        return response


@app.route('/admin/leave/action/<int:leave_id>', methods=['POST'])
@login_required
@admin_required
def leave_action(leave_id):
    leave = Leave.query.get_or_404(leave_id)
    action = request.form.get('action')
    comment = request.form.get('comment')

    if action == 'approve':
        leave.status = 'Approved'
        leave.approved_by = current_user.id
        leave.approved_date = datetime.utcnow()
    elif action == 'reject':
        leave.status = 'Rejected'
        leave.approved_by = current_user.id
        leave.approved_date = datetime.utcnow()

    leave.admin_comment = comment
    db.session.commit()

    flash(f'Leave {action}d successfully!', 'success')
    return redirect(url_for('admin_leaves'))


@app.route('/admin/leaves')
@login_required
@admin_required
def admin_leaves():
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    leave_type_filter = request.args.get('leave_type')
    department_filter = request.args.get('department')
    date_from_filter = request.args.get('date_from')
    search_filter = request.args.get('search')

    # Build query
    query = Leave.query.join(User, Leave.user_id == User.id)

    # Apply filters
    if status_filter != 'all':
        query = query.filter(Leave.status == status_filter)

    if leave_type_filter and leave_type_filter != '':
        query = query.filter(Leave.leave_type == leave_type_filter)

    if department_filter and department_filter != '':
        query = query.filter(User.department == department_filter)

    if date_from_filter and date_from_filter != '':
        try:
            filter_date = datetime.strptime(date_from_filter, '%Y-%m-%d').date()
            query = query.filter(Leave.start_date >= filter_date)
        except ValueError:
            pass  # Ignore invalid date format

    if search_filter and search_filter != '':
        search = f'%{search_filter}%'
        query = query.filter(
            db.or_(
                User.first_name.ilike(search),
                User.last_name.ilike(search),
                User.employee_id.ilike(search),
                User.designation.ilike(search),
                User.phone.ilike(search)
            )
        )

    # Order by most recent first
    leaves = query.order_by(Leave.applied_date.desc()).all()

    return render_template('admin/leaves.html',
                           leaves=leaves,
                           status_filter=status_filter,
                           leave_type_filter=leave_type_filter,
                           department_filter=department_filter,
                           date_from_filter=date_from_filter,
                           search_filter=search_filter)

@app.route('/admin/employees')
@login_required
@admin_required
def admin_employees():
    employees = User.query.filter_by(is_admin=False).all()
    return render_template('admin/employees.html', employees=employees)


@app.route('/api/employee/<int:id>/deactivate', methods=['POST'])
@login_required
@admin_required
def deactivate_employee(id):
    employee = User.query.get_or_404(id)
    if employee.is_admin:
        return jsonify({'error': 'Cannot deactivate admin users'}), 400

    employee.is_active = False
    db.session.commit()
    return jsonify({'message': 'Employee deactivated successfully'})
@app.route('/api/employee/<int:id>/activate', methods=['POST'])
@login_required
@admin_required
def activate_employee(id):
    employee = User.query.get_or_404(id)
    employee.is_active = True
    db.session.commit()
    return jsonify({'message': 'Employee activated successfully'})


@app.route('/api/employee/<int:id>/update', methods=['POST'])
@login_required
@admin_required
def update_employee(id):
    employee = User.query.get_or_404(id)

    # Update employee details
    employee.first_name = request.form.get('first_name', employee.first_name)
    employee.last_name = request.form.get('last_name', employee.last_name)
    employee.email = request.form.get('email', employee.email)
    employee.phone = request.form.get('phone', employee.phone)
    employee.department = request.form.get('department', employee.department)
    employee.designation = request.form.get('designation', employee.designation)

    # Handle checkbox for active status
    is_active = request.form.get('is_active') == 'on'
    employee.is_active = is_active

    db.session.commit()
    return jsonify({'message': 'Employee updated successfully'})

@app.route('/admin/attendance')
@login_required
@admin_required
def admin_attendance():
    date_filter = request.args.get('date', date.today().isoformat())

    try:
        filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
    except ValueError:
        filter_date = date.today()

    # Get attendance records for the date with employee details
    attendance_records = Attendance.query.options(joinedload(Attendance.employee)).filter_by(date=filter_date).all()

    # Get all employees for the dropdown
    employees = User.query.filter_by(is_admin=False, is_active=True).all()

    return render_template('admin/attendance.html',
                           attendance_records=attendance_records,
                           date_filter=date_filter,
                           employees=employees)


from flask import make_response, request, jsonify, send_file
import csv
import io
from datetime import date, timedelta
import json
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
import os


@app.route('/admin/reports', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_reports():
    from datetime import date

    # Monthly attendance summary
    current_month = date.today().month
    current_year = date.today().year

    # Get attendance summary for current month
    attendance_summary = db.session.query(
        Attendance.status,
        db.func.count(Attendance.id).label('count')
    ).filter(
        db.extract('year', Attendance.date) == current_year,
        db.extract('month', Attendance.date) == current_month
    ).group_by(Attendance.status).all()

    # Get leave summary by type
    leave_summary = db.session.query(
        Leave.leave_type,
        db.func.count(Leave.id).label('count'),
        db.func.sum(Leave.total_days).label('total_days')
    ).filter(
        Leave.status == 'Approved'
    ).group_by(Leave.leave_type).all()

    # Get department-wise employee count
    dept_summary = db.session.query(
        User.department,
        db.func.count(User.id).label('count')
    ).filter(
        User.is_admin == False,
        User.is_active == True
    ).group_by(User.department).all()

    # Monthly attendance trend (last 6 months)
    monthly_trend = []
    for i in range(5, -1, -1):
        month_date = date.today() - timedelta(days=30 * i)
        month = month_date.month
        year = month_date.year

        present_count = Attendance.query.filter(
            db.extract('year', Attendance.date) == year,
            db.extract('month', Attendance.date) == month,
            Attendance.status == 'Present'
        ).count()

        total_employees = User.query.filter_by(is_admin=False, is_active=True).count()
        attendance_rate = (present_count / (total_employees * 22 * 0.01)) if total_employees > 0 else 0

        monthly_trend.append({
            'month': month_date.strftime('%b %Y'),
            'present': present_count,
            'rate': min(attendance_rate, 100)  # Cap at 100%
        })

    # Calculate summary statistics
    total_employees = User.query.filter_by(is_admin=False, is_active=True).count()
    present_days = Attendance.query.filter(
        db.extract('year', Attendance.date) == current_year,
        db.extract('month', Attendance.date) == current_month,
        Attendance.status == 'Present'
    ).count()
    approved_leaves = Leave.query.filter_by(status='Approved').count()

    # Calculate attendance rate
    attendance_rate = (present_days / (total_employees * 22 * 0.01)) if total_employees > 0 else 0

    # Handle POST request for export
    if request.method == 'POST':
        format_type = request.form.get('format')

        if format_type == 'csv':
            return export_csv(attendance_summary, leave_summary, dept_summary, monthly_trend,
                              total_employees, present_days, approved_leaves, attendance_rate)
        elif format_type == 'excel':
            return export_excel(attendance_summary, leave_summary, dept_summary, monthly_trend,
                                total_employees, present_days, approved_leaves, attendance_rate)
        elif format_type == 'json':
            return export_json(attendance_summary, leave_summary, dept_summary, monthly_trend,
                               total_employees, present_days, approved_leaves, attendance_rate)
        elif format_type == 'pdf':
            return export_pdf(attendance_summary, leave_summary, dept_summary, monthly_trend,
                              total_employees, present_days, approved_leaves, attendance_rate)
        else:
            return jsonify({'error': 'Invalid format specified'}), 400

    # For GET requests, render the template
    return render_template('admin/reports.html',
                           attendance_summary=attendance_summary,
                           leave_summary=leave_summary,
                           dept_summary=dept_summary,
                           monthly_trend=monthly_trend,
                           current_month=current_month,
                           current_year=current_year,
                           today=date.today(),
                           total_employees=total_employees,
                           present_days=present_days,
                           approved_leaves=approved_leaves,
                           attendance_rate=min(attendance_rate, 100))


def export_csv(attendance_summary, leave_summary, dept_summary, monthly_trend,
               total_employees, present_days, approved_leaves, attendance_rate):
    """Export data as CSV file"""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['TextileLeave Pro - System Report'])
    writer.writerow(['Generated on:', date.today().strftime('%Y-%m-%d')])
    writer.writerow([])

    # Summary Statistics
    writer.writerow(['SUMMARY STATISTICS'])
    writer.writerow(['Total Employees', total_employees])
    writer.writerow(['Present Days (Current Month)', present_days])
    writer.writerow(['Approved Leaves', approved_leaves])
    writer.writerow(['Attendance Rate', f"{min(attendance_rate, 100):.1f}%"])
    writer.writerow([])

    # Attendance Summary
    writer.writerow(['ATTENDANCE SUMMARY (Current Month)'])
    writer.writerow(['Status', 'Count', 'Percentage'])
    total_attendance = sum(count for _, count in attendance_summary)
    for status, count in attendance_summary:
        percentage = (count / total_attendance * 100) if total_attendance > 0 else 0
        writer.writerow([status, count, f"{percentage:.1f}%"])
    writer.writerow([])

    # Leave Summary
    writer.writerow(['LEAVE SUMMARY'])
    writer.writerow(['Leave Type', 'Applications', 'Total Days', 'Average Duration'])
    for leave_type, count, total_days in leave_summary:
        avg_duration = (total_days / count) if count > 0 else 0
        writer.writerow([leave_type, count, total_days, f"{avg_duration:.1f} days"])
    writer.writerow([])

    # Department Distribution
    writer.writerow(['DEPARTMENT DISTRIBUTION'])
    writer.writerow(['Department', 'Employees', 'Percentage'])
    total_dept = sum(count for _, count in dept_summary)
    for department, count in dept_summary:
        percentage = (count / total_dept * 100) if total_dept > 0 else 0
        dept_name = department or 'Not Specified'
        writer.writerow([dept_name, count, f"{percentage:.1f}%"])
    writer.writerow([])

    # Attendance Trend
    writer.writerow(['ATTENDANCE TREND (Last 6 Months)'])
    writer.writerow(['Month', 'Present Days', 'Attendance Rate'])
    for month_data in monthly_trend:
        writer.writerow([month_data['month'], month_data['present'], f"{month_data['rate']:.1f}%"])

    # Create response
    response = make_response(output.getvalue())
    response.headers[
        'Content-Disposition'] = f'attachment; filename=textileleave_report_{date.today().strftime("%Y%m%d")}.csv'
    response.headers['Content-type'] = 'text/csv'
    return response


def export_excel(attendance_summary, leave_summary, dept_summary, monthly_trend,
                 total_employees, present_days, approved_leaves, attendance_rate):
    """Export data as Excel file"""
    # Create a new workbook
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(['TextileLeave Pro - System Report'])
    ws_summary.append(['Generated on:', date.today().strftime('%Y-%m-%d')])
    ws_summary.append([])
    ws_summary.append(['SUMMARY STATISTICS'])
    ws_summary.append(['Total Employees', total_employees])
    ws_summary.append(['Present Days (Current Month)', present_days])
    ws_summary.append(['Approved Leaves', approved_leaves])
    ws_summary.append(['Attendance Rate', f"{min(attendance_rate, 100):.1f}%"])

    # Attendance sheet
    ws_attendance = wb.create_sheet(title="Attendance")
    ws_attendance.append(['ATTENDANCE SUMMARY'])
    ws_attendance.append(['Status', 'Count', 'Percentage'])
    total_attendance = sum(count for _, count in attendance_summary)
    for status, count in attendance_summary:
        percentage = (count / total_attendance * 100) if total_attendance > 0 else 0
        ws_attendance.append([status, count, f"{percentage:.1f}%"])

    # Leave sheet
    ws_leave = wb.create_sheet(title="Leaves")
    ws_leave.append(['LEAVE SUMMARY'])
    ws_leave.append(['Leave Type', 'Applications', 'Total Days', 'Average Duration'])
    for leave_type, count, total_days in leave_summary:
        avg_duration = (total_days / count) if count > 0 else 0
        ws_leave.append([leave_type, count, total_days, f"{avg_duration:.1f}"])

    # Department sheet
    ws_dept = wb.create_sheet(title="Departments")
    ws_dept.append(['DEPARTMENT DISTRIBUTION'])
    ws_dept.append(['Department', 'Employees', 'Percentage'])
    total_dept = sum(count for _, count in dept_summary)
    for department, count in dept_summary:
        percentage = (count / total_dept * 100) if total_dept > 0 else 0
        dept_name = department or 'Not Specified'
        ws_dept.append([dept_name, count, f"{percentage:.1f}%"])

    # Trend sheet
    ws_trend = wb.create_sheet(title="Trend")
    ws_trend.append(['ATTENDANCE TREND'])
    ws_trend.append(['Month', 'Present Days', 'Attendance Rate'])
    for month_data in monthly_trend:
        ws_trend.append([month_data['month'], month_data['present'], f"{month_data['rate']:.1f}%"])

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'textileleave_report_{date.today().strftime("%Y%m%d")}.xlsx'
    )


def export_json(attendance_summary, leave_summary, dept_summary, monthly_trend,
                total_employees, present_days, approved_leaves, attendance_rate):
    """Export data as JSON file"""
    report_data = {
        'report_date': date.today().isoformat(),
        'summary': {
            'total_employees': total_employees,
            'present_days': present_days,
            'approved_leaves': approved_leaves,
            'attendance_rate': min(attendance_rate, 100)
        },
        'attendance_summary': [
            {'status': status, 'count': count}
            for status, count in attendance_summary
        ],
        'leave_summary': [
            {
                'leave_type': leave_type,
                'applications': count,
                'total_days': total_days,
                'average_duration': (total_days / count) if count > 0 else 0
            }
            for leave_type, count, total_days in leave_summary
        ],
        'department_distribution': [
            {
                'department': department or 'Not Specified',
                'employees': count
            }
            for department, count in dept_summary
        ],
        'attendance_trend': monthly_trend
    }

    json_data = json.dumps(report_data, indent=2, default=str)
    response = make_response(json_data)
    response.headers[
        'Content-Disposition'] = f'attachment; filename=textileleave_report_{date.today().strftime("%Y%m%d")}.json'
    response.headers['Content-type'] = 'application/json'
    return response


def export_pdf(attendance_summary, leave_summary, dept_summary, monthly_trend,
               total_employees, present_days, approved_leaves, attendance_rate):
    """Export data as PDF file"""
    try:
        # Try to import reportlab if available
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        from reportlab.lib.units import inch

        # Create PDF in memory
        pdf_buffer = BytesIO()

        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph(f"<b>TextileLeave Pro - System Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Paragraph(f"Generated on: {date.today().strftime('%Y-%m-%d')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        elements.append(Paragraph("<b>SUMMARY STATISTICS</b>", styles['Heading2']))
        summary_data = [
            ['Total Employees', str(total_employees)],
            ['Present Days (Current Month)', str(present_days)],
            ['Approved Leaves', str(approved_leaves)],
            ['Attendance Rate', f"{min(attendance_rate, 100):.1f}%"]
        ]
        summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))

        # Build the PDF
        doc.build(elements)
        pdf_buffer.seek(0)

        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'textileleave_report_{date.today().strftime("%Y%m%d")}.pdf'
        )

    except ImportError:
        # If reportlab is not installed, return a simple text file
        text_content = f"""TextileLeave Pro - System Report
Generated on: {date.today().strftime('%Y-%m-%d')}

SUMMARY STATISTICS:
Total Employees: {total_employees}
Present Days: {present_days}
Approved Leaves: {approved_leaves}
Attendance Rate: {min(attendance_rate, 100):.1f}%

Note: PDF export requires reportlab library.
Install with: pip install reportlab
"""
        response = make_response(text_content)
        response.headers[
            'Content-Disposition'] = f'attachment; filename=textileleave_report_{date.today().strftime("%Y%m%d")}.txt'
        response.headers['Content-type'] = 'text/plain'
        return response


@app.route('/admin/report', methods=['POST'])
@login_required
@admin_required
def export_report():
    """Alternative route for exporting reports"""
    format_type = request.form.get('format', 'csv')

    # Get fresh data for export
    current_month = date.today().month
    current_year = date.today().year

    # Get attendance summary for current month
    attendance_summary = db.session.query(
        Attendance.status,
        db.func.count(Attendance.id).label('count')
    ).filter(
        db.extract('year', Attendance.date) == current_year,
        db.extract('month', Attendance.date) == current_month
    ).group_by(Attendance.status).all()

    # Get other data similarly...
    # (Add the same data fetching logic as in admin_reports)

    # Call appropriate export function based on format
    if format_type == 'excel':
        return export_excel(attendance_summary, [], [], [], 0, 0, 0, 0)
    elif format_type == 'json':
        return export_json(attendance_summary, [], [], [], 0, 0, 0, 0)
    elif format_type == 'pdf':
        return export_pdf(attendance_summary, [], [], [], 0, 0, 0, 0)
    else:
        return export_csv(attendance_summary, [], [], [], 0, 0, 0, 0)
@app.route('/mark_attendance', methods=['POST'])
@login_required
@admin_required
def mark_attendance():
    employee_id = request.form.get('employee_id')
    date_str = request.form.get('date')
    check_in = request.form.get('check_in')
    check_out = request.form.get('check_out')
    status = request.form.get('status')
    remarks = request.form.get('remarks')

    try:
        attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Invalid date format!', 'danger')
        return redirect(url_for('admin_attendance'))

    # Check if employee exists
    employee = User.query.get(employee_id)
    if not employee:
        flash('Employee not found!', 'danger')
        return redirect(url_for('admin_attendance'))

    # Check for existing attendance record
    attendance = Attendance.query.filter_by(
        user_id=employee_id,
        date=attendance_date
    ).first()

    if not attendance:
        attendance = Attendance(
            user_id=employee_id,
            date=attendance_date,
            recorded_by=current_user.id
        )

    if check_in:
        try:
            attendance.check_in = datetime.strptime(check_in, '%H:%M').time()
        except ValueError:
            attendance.check_in = None

    if check_out:
        try:
            attendance.check_out = datetime.strptime(check_out, '%H:%M').time()
        except ValueError:
            attendance.check_out = None

    attendance.status = status
    attendance.remarks = remarks

    db.session.add(attendance)
    db.session.commit()

    flash('Attendance marked successfully!', 'success')
    return redirect(url_for('admin_attendance', date=date_str))

@app.route('/user/profile')
@login_required
def user_profile():
    # Calculate leave statistics for the user
    leaves = Leave.query.filter_by(user_id=current_user.id).all()

    total_leaves = len(leaves)
    approved_leaves = len([l for l in leaves if l.status == 'Approved'])
    pending_leaves = len([l for l in leaves if l.status == 'Pending'])
    rejected_leaves = len([l for l in leaves if l.status == 'Rejected'])

    return render_template('user/profile.html',
                           total_leaves=total_leaves,
                           approved_leaves=approved_leaves,
                           pending_leaves=pending_leaves,
                           rejected_leaves=rejected_leaves)

@app.route('/user/attendance')
@login_required
def user_attendance():
    month = request.args.get('month', date.today().month)
    year = request.args.get('year', date.today().year)

    try:
        month = int(month)
        year = int(year)
    except ValueError:
        month = date.today().month
        year = date.today().year

    attendances = Attendance.query.filter_by(user_id=current_user.id) \
        .filter(db.extract('year', Attendance.date) == year) \
        .filter(db.extract('month', Attendance.date) == month) \
        .order_by(Attendance.date).all()

    # Calculate statistics
    present_count = len([a for a in attendances if a.status == 'Present'])
    absent_count = len([a for a in attendances if a.status == 'Absent'])
    late_count = len([a for a in attendances if a.status == 'Late'])
    half_day_count = len([a for a in attendances if a.status == 'Half-day'])

    return render_template('user/attendance.html',
                           attendances=attendances,
                           month=month,
                           year=year,
                           present_count=present_count,
                           absent_count=absent_count,
                           late_count=late_count,
                           half_day_count=half_day_count)



# Initialize database
def init_db():
    with app.app_context():
        try:
            # Create tables
            db.create_all()
            print("✅ Database tables created successfully!")

            # Check if admin exists
            admin_email = 'admin@textile.com'
            admin_user = User.query.filter_by(email=admin_email).first()

            if not admin_user:
                admin = User(
                    employee_id='ADMIN001',
                    first_name='Admin',
                    last_name='User',
                    email=admin_email,
                    phone='+1234567890',
                    department='Administration',
                    designation='System Administrator',
                    date_of_joining=date.today(),
                    is_admin=True,
                    is_active=True
                )
                admin.password = 'admin123'  # Change this in production!
                db.session.add(admin)
                db.session.commit()
                print("✅ Admin user created successfully!")
                print("📧 Email: admin@textile.com")
                print("🔑 Password: admin123")

            # Create a test employee
            test_email = 'employee@textile.com'
            test_user = User.query.filter_by(email=test_email).first()

            if not test_user:
                employee = User(
                    employee_id='EMP001',
                    first_name='John',
                    last_name='Doe',
                    email=test_email,
                    phone='+1234567891',
                    department='Production',
                    designation='Operator',
                    date_of_joining=date.today(),
                    is_admin=False,
                    is_active=True
                )
                employee.password = 'employee123'
                db.session.add(employee)
                db.session.commit()
                print("✅ Test employee created successfully!")
                print("📧 Email: employee@textile.com")
                print("🔑 Password: employee123")

        except Exception as e:
            print(f"❌ Database initialization failed: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    print("🚀 Starting Textile LMS Application...")
    print(f"📅 Current date: {date.today()}")
    print(f"🏠 Running on: http://127.0.0.1:5000")

    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)